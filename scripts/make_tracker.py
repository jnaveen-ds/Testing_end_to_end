"""Generate docs/Learning_Tracker.xlsx — the day-by-day Azure learning tracker.

Run from repo root:  .venv/bin/python scripts/make_tracker.py
"""
import datetime as dt

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation

HDR = PatternFill("solid", fgColor="1F4E78")
HDR_FONT = Font(bold=True, color="FFFFFF", size=10)
WRAP = Alignment(wrap_text=True, vertical="top")

# (day, date, phase, goal, azure services, portal steps, CLI commands, time)
DAYS = [
    (1, "2026-08-31", "Setup", "Repo governance + OIDC prep + first local run",
     "none",
     "Branch protection on main; GHCR packages public; Entra app + federated credential; Contributor role",
     "az login; az account list; docker compose up --build (local)", "1h"),
    (2, "2026-09-01", "Foundations", "Azure anatomy: subscription, RGs, providers",
     "Resource Groups",
     "Subscriptions, Resource groups, Cost Management blades",
     "az login; az account set; az group create/delete; az resource list", "1h"),
    (3, "2026-09-02", "Secrets", "Key Vault manually in the portal",
     "Key Vault, RBAC",
     "Create vault (RBAC); Secrets tab add LLM-PROVIDER; IAM role assignment",
     "az keyvault show; az keyvault secret set/show", "1h"),
    (4, "2026-09-03", "Secrets", "CLI-driven vault access with least privilege",
     "Key Vault, Entra ID",
     "Role assignments blade on the vault",
     "az ad sp create; az role assignment create; az keyvault secret show", "1h"),
    (5, "2026-09-04", "IaC", "Terraform 101: resource group via code; state concept",
     "Terraform (azurerm)",
     "Verify TF-created RG in portal; read the plan output line by line",
     "terraform init; plan -out; apply; state list; destroy", "1h"),
    (6, "2026-09-04", "IaC", "Key Vault + secret entirely in Terraform",
     "Key Vault via TF",
     "Compare portal form fields with TF resource attributes",
     "terraform apply; terraform destroy", "1h"),
    (7, "2026-09-04", "Integration", "App reads config from Key Vault; OIDC app ready for GitHub",
     "Key Vault, Entra federated credential",
     "Federated credential subject: repo:jnaveen-ds/Testing_end_to_end:ref:refs/heads/main",
     "az keyvault secret show; verify app /health uses vault config", "1h"),
    (8, "2026-09-06", "Cleanup", "Week 1 review + cost check",
     "Cost Management",
     "Cost analysis blade; budget review",
     "az group list; az consumption usage list", "30m"),
    (9, "2026-09-07", "Week 2 · VM", "Terraform: Linux VM + VNet + subnet + NSG + public IP",
     "Linux VM, VNet, NSG, disk, public IP",
     "VM blade: boot diagnostics, NSG rules, size/cost estimate",
     "terraform apply (infra/day9); az vm list; ssh", "1h"),
    (10, "2026-09-08", "VM", "Deploy the app stack onto the VM; open only needed ports",
     "VM, Docker Compose, NSG rules",
     "NSG inbound rules audit; boot diagnostics",
     "ssh; docker compose up -d --build; curl /health", "90m"),
    (11, "2026-09-09", "VM", "TLS + reverse proxy in front of the app",
     "nginx, TLS (certbot)",
     "Certificates; custom domain (if any)",
     "nginx config; certbot --nginx or openssl self-signed", "1h"),
    (12, "2026-09-11", "CD", "Actions deploys to the VM: SSH + pull sha-tagged image + restart",
     "GitHub Actions, VM",
     "GitHub environment secrets review (never in git)",
     "git push -> deploy job -> curl public /health", "1h"),
    (13, "2026-09-11", "Chaos", "Break things on purpose: kill worker, wrong env, full disk; recover",
     "ops skills",
     "Serial console; boot diagnostics",
     "systemctl; docker; df -h; journalctl", "90m"),
    (14, "2026-09-13", "Cleanup", "Destroy the VM resource group; review week-2 spend",
     "Cost Management",
     "Cost analysis", "az group delete; terraform destroy", "30m"),
    (15, "2026-09-14", "Week 3 · Containers", "Log Analytics + Container Apps environment + deploy API",
     "Log Analytics, Container Apps",
     "Container app blade: Revisions, Log stream, Scale",
     "az containerapp env create; TF apply (infra/day15)", "90m"),
    (16, "2026-09-15", "Containers", "Worker as second app; scale-to-zero via KEDA rule",
     "Container Apps scale rules (KEDA)",
     "Scale tab: min 0 / max 3; watch 0->1 on submit",
     "az containerapp update --min-replicas 0", "1h"),
    (17, "2026-09-16", "Data", "Azure PostgreSQL Flexible (B1ms) wired via Key Vault secret",
     "PostgreSQL Flexible, Key Vault",
     "Server networking + metrics blades; firewall rules",
     "TF apply; az keyvault secret set (connection string)", "90m"),
    (18, "2026-09-17", "Autoscaling", "Load test with hey; watch replicas scale in the Metrics blade",
     "Container Apps autoscale, Metrics",
     "Metrics blade: replica count vs request rate",
     "hey -c 20 -z 2m https://<app-url>", "1h"),
    (19, "2026-09-19", "Website", "WEBSITE DEPLOYMENT: React UI on Azure Static Web Apps (free tier), pointed at the API",
     "Static Web Apps (+ optional custom domain)",
     "SWA portal: environments, tokens, custom domain",
     "az staticwebapp create; swa deploy; verify site + API wiring", "90m"),
    (20, "2026-09-20", "Resilience", "Rollback drills: bad revision -> revert; traffic split blue/green",
     "Container Apps revisions",
     "Revision management + traffic weighting UI",
     "az containerapp revision list/activate; ingress traffic set", "1h"),
    (21, "2026-09-20", "Cleanup", "Destroy week-3 compute; keep only free-tier DB + images",
     "-", "Cost check", "az group delete (each learn RG)", "30m"),
    (22, "2026-09-22", "Kubernetes", "AKS evening: 1-node free-tier cluster; deploy API+worker; DESTROY tonight",
     "AKS, kubectl, Azure Load Balancer",
     "AKS blade: workloads, services, logs",
     "az aks create; get-credentials; kubectl apply/get; az aks delete", "2h"),
    (23, "2026-09-23", "Config", "App Configuration + feature flag; change behavior without redeploy",
     "App Configuration",
     "Feature manager UI",
     "az appconfig create; az appconfig kv set; wire app endpoint", "1h"),
    (24, "2026-09-24", "Resilience", "App Insights live; kill DB mid-traffic; alert on error rate",
     "Application Insights, Monitor alerts",
     "Live Metrics; Log Analytics queries; create alert rule",
     "az monitor metrics alert create", "90m"),
    (25, "2026-09-25", "Pipeline", "Full automated flow: PR -> CI -> publish -> deploy staging -> approval -> prod -> rollback",
     "OIDC deploys, environments, approvals",
     "GitHub environment protection rules",
     "watch the Actions run end to end", "90m"),
    (26, "2026-09-26", "FinOps", "Sweep + destroy everything not deliberately kept; review spend",
     "Cost Management",
     "Cost analysis vs budget; verify alerts fired",
     "az group list; az group delete per learn-* RG", "90m"),
    (27, "2026-09-27", "Portfolio", "Screenshot dashboards -> docs/PORTFOLIO.md -> destroy the rest",
     "-", "Screenshot everything first", "final az group delete", "1h"),
]

DEPLOYMENTS = [
    (1, "Local compose stack (5 containers)", "Docker Compose", "Day 1", "keep (local)"),
    (2, "Key Vault-backed app configuration", "Key Vault + RBAC", "Day 7", "vault may stay (~$0)"),
    (3, "Full stack on a VM + TLS + CI deploy-on-push", "VM, VNet, NSG, nginx, TLS", "Days 9-12", "destroy Day 14"),
    (4, "Container Apps: API + scale-to-zero worker", "Container Apps, KEDA, Log Analytics", "Day 17", "destroy Day 21"),
    (5, "Managed PostgreSQL wired through Key Vault", "PostgreSQL Flexible", "Day 17", "destroy Day 21 (or keep within free tier)"),
    (6, "WEBSITE: React UI on Azure Static Web Apps", "Static Web Apps", "Day 19", "destroy Day 26"),
    (7, "Rollback + blue/green traffic split drill", "Container Apps revisions", "Day 20", "-"),
    (8, "AKS: 1-node cluster, deploy, destroy same evening", "AKS, kubectl", "Day 22", "destroy same night"),
    (9, "Load-tested + monitored deployment with alerts", "App Insights, Monitor", "Day 24", "destroy Day 26"),
    (10, "Full CI/CD: PR -> test -> publish -> deploy -> approve -> rollback", "Actions + OIDC", "Day 25", "-"),
]


def header(ws, cols):
    for i, name in enumerate(cols, 1):
        c = ws.cell(row=1, column=i, value=name)
        c.fill = HDR
        c.font = HDR_FONT


def set_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w


def main():
    wb = Workbook()

    # ---------- Overview ----------
    ws = wb.active
    ws.title = "Overview"
    lines = [
        ("Azure / DevOps Learning Tracker", ""),
        ("App", "Feedback Analyzer (FastAPI + Celery + React) - one app, many deployments"),
        ("Window", "Aug 31 - Sep 27 2026, ~1 h/day (Day 22 AKS = 2 h)"),
        ("Budget", "$200 trial credit; expected spend $20-60; destroy same-day"),
        ("Golden rule", "A day is Done only when Destroyed? = Yes (except keep-listed items)"),
        ("Your role", "Portal clicks, az commands, terraform plan review + apply, verification"),
        ("Agent's job", "TF files + exact commands prepared before each session; docs updated"),
        ("Docs", "docs/LEARNING_PLAN.md · docs/DAILY_PLAYBOOK.md · docs/RUNBOOK.md"),
        ("Website deployment", "Day 19 (Fri Sep 19): React UI on Azure Static Web Apps, its own URL"),
        ("Deployments target", "10 tracked deployments - see 'Deployments' tab"),
    ]
    for i, (a, b) in enumerate(lines, 1):
        ca = ws.cell(row=i, column=1, value=a)
        cb = ws.cell(row=i, column=2, value=b)
        cb.alignment = WRAP
        ca.font = Font(bold=True, size=14 if i == 1 else 10)
        cb.font = Font(size=10)
    set_widths(ws, [26, 100])

    # ---- Daily Plan ----
    ws = wb.create_sheet("Daily Plan")
    header(ws, ["Day", "Date", "Dow", "Phase", "Goal", "Azure services",
                "Portal steps", "CLI commands (key)", "Time", "Status", "Cost $", "Destroyed?", "Notes"])
    for r, (day, date, phase, goal, services, portal, cli, t) in enumerate(DAYS, 2):
        d = dt.date.fromisoformat(date)
        vals = [day, date, d.strftime("%a"), phase, goal, services, portal, cli, t,
                "Not started", "", "", ""]
        for c, v in enumerate(vals, 1):
            cell = ws.cell(row=r, column=c, value=v)
            cell.font = Font(size=10)
            cell.alignment = WRAP
    dv = DataValidation(type="list", formula1='"Not started,In progress,Done,Skipped"', allow_blank=True)
    ws.add_data_validation(dv)
    dv.add(f"J2:J{len(DAYS) + 1}")
    dv2 = DataValidation(type="list", formula1='"Yes,No,-"', allow_blank=True)
    ws.add_data_validation(dv2)
    dv2.add(f"M2:M{len(DAYS) + 1}")
    set_widths(ws, [5, 11, 6, 12, 34, 24, 38, 38, 6, 12, 8, 11, 22])
    ws.freeze_panes = "E2"

    # ---- Deployments ----
    ws = wb.create_sheet("Deployments")
    header(ws, ["#", "Deployment", "Azure services", "Day", "End-state plan", "Status", "Evidence link"])
    for r, (n, what, services, day, plan) in enumerate(DEPLOYMENTS, 2):
        for c, v in enumerate([n, what, services, day, plan, "Not started", ""], 1):
            cell = ws.cell(row=r, column=c, value=v)
            cell.font = Font(size=10)
            cell.alignment = WRAP
    set_widths(ws, [4, 44, 28, 10, 30, 12, 30])

    # ---- Cost Log ----
    ws = wb.create_sheet("Cost Log")
    header(ws, ["Date", "Resource group", "Service", "Est. cost $", "Actual cost $", "Destroyed?", "Notes"])
    ws.cell(row=40, column=5, value="TOTAL:").font = Font(bold=True)
    ws.cell(row=40, column=6, value="=SUM(E2:E39)").font = Font(bold=True)
    set_widths(ws, [12, 26, 26, 12, 12, 11, 30])

    # ---- Destroy Ritual ----
    ws = wb.create_sheet("Destroy Ritual")
    steps = [
        "1. SCREENSHOT dashboards first (portfolio evidence)",
        "2. az group list -o table   -- know exactly what exists",
        "3. az group delete -n learn-<stage>-<date> --yes --no-wait",
        "4. Terraform days: cd infra/<day> && terraform destroy  -- review: must list only today's resources",
        "5. az resource list -o table | grep <date>   -- no orphaned disks/NICs/IPs",
        "6. az consumption usage list --top 5 -o table   -- know today's cost",
        "7. Next morning: az group list   -- nothing unexpected alive",
    ]
    ws.cell(row=1, column=1, value="Run at the END of EVERY session. No exceptions.").font = Font(bold=True, size=12)
    for i, s in enumerate(steps, 3):
        ws.cell(row=i, column=1, value=s).font = Font(size=10)
    set_widths(ws, [110])

    wb.save("docs/Learning_Tracker.xlsx")
    print("saved docs/Learning_Tracker.xlsx")


if __name__ == "__main__":
    main()
